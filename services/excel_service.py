"""
Generates the downloadable Excel (.xlsx) report from order data.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("no", "No"),
    ("date", "Order Date"),
    ("customer_name", "Customer Name"),
    ("order_id", "Order ID"),
    ("tracking_id", "Tracking ID"),
    ("sku", "SKU"),
    ("quantity", "Quantity"),
    ("selling_price", "Selling Price"),
    ("purchase_cost", "Cost Price"),
    ("status", "Status"),
    ("payment_status", "Payment Status"),
    ("profit", "Net Profit"),
]

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

HEADER_FILL = PatternFill(start_color="0F2A4A", end_color="0F2A4A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(*(Side(style="thin", color="DDDDDD"),) * 4)


def generate_excel_report(orders, report_name=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # header
    for col_idx, (key, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, order in enumerate(orders, start=2):
        for col_idx, (key, _label) in enumerate(COLUMNS, start=1):
            if key == "no":
                value = row_idx - 1
            else:
                value = order.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if key in ("selling_price", "purchase_cost", "payment_status", "profit"):

                cell.number_format = "#,##0.00"
            if key == "profit":
                if isinstance(value, (int, float)):
                    if value < 0:
                        cell.font = Font(color="C0392B", bold=True)
                    elif value == 0:
                        cell.font = Font(color="D68910", bold=True)
                    else:
                        cell.font = Font(color="1E8449", bold=True)
            if key == "margin" and isinstance(value, (int, float)):
                cell.number_format = "0.00"

    # column widths
    widths = [8, 14, 24, 24, 20, 16, 10, 16, 14, 14, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # summary sheet
    ws2 = wb.create_sheet("Summary")
    revenue = round(sum(o["revenue"] for o in orders), 2)
    total_cost = round(sum(o["total_cost"] for o in orders), 2)
    profit = round(sum(float(o.get("profit", 0)) for o in orders), 2)
    margin = round((profit / revenue) * 100, 2) if revenue else 0
    summary_rows = [
        ("Report Generated", datetime.now().strftime("%d %b %Y, %I:%M %p")),
        ("Total Orders", len(orders)),
        ("Total Revenue (₹)", revenue),
        ("Total Cost (₹)", total_cost),
        ("Total Profit (₹)", profit),
        ("Profit Margin (%)", margin),
    ]
    for i, (label, value) in enumerate(summary_rows, start=1):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22

    if not report_name:
        report_name = f"Sales_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    filename = f"{report_name}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)
    return filepath, filename